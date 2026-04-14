from __future__ import annotations

import html
import json
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config.entities import INCLUDE_BANKS_LAYER
from config.paths import DIST_DIR, EXCEL_OUTPUT, HTML_INDEX, SCORES_CSV, SCORING_AUDIT_DIR
from config.topics import TOPICS
from src.search import SEARCH_DATE_LABEL
from src.utils import console, setup_logger

logger = setup_logger(__name__)

TOPIC_BY_NAME: dict[str, dict[str, str | list[str]]] = {str(t["topic_name"]): t for t in TOPICS}

HEADER_FILL = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000", size=11)
BODY_FONT = Font(color="000000", size=10)
BODY_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")


def _topic_display_name(topic_key: str) -> str:
    """Human-readable topic name for Excel headers (matches dashboard short_label when set)."""
    meta = TOPIC_BY_NAME.get(topic_key)
    if meta and meta.get("short_label"):
        return str(meta["short_label"])
    return topic_key.replace("_", " ").title()


def _build_topic_description(topic_row: dict[str, str | list[str]]) -> str:
    """Layman description for theme card inline text."""
    return str(topic_row["layman_description"])


def _layer_display_name(layer: str) -> str:
    """Display layer as Lender, Borrower or Bank."""
    return layer.title() if layer else layer


def _style_header(ws: Any) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _style_body(ws: Any, n_rows: int, n_cols: int) -> None:
    for row_idx in range(2, n_rows + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = BODY_FILL
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")


def _auto_width(ws: Any) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_excel(df: pd.DataFrame) -> Path:
    """Generate the multi-sheet Excel workbook from scored DataFrame."""
    wb = Workbook()

    # ── Sheet 1: Lender Ranking ──────────────────────────────────────────────
    ws_lender = wb.active
    ws_lender.title = "Lender Ranking"
    lender_df = (
        df[df["layer"] == "lender"]
        .sort_values("terms_power_score", ascending=False)
        .reset_index(drop=True)
    )
    lender_cols = [
        "entity_name", "ticker", "terms_power_score",
        "positive_count", "negative_count",
        "top_negative_topic", "top_positive_topic",
    ]
    _write_ranked_sheet(ws_lender, lender_df, lender_cols, score_col="terms_power_score")

    # ── Sheet 2: Borrower Distress ───────────────────────────────────────────
    ws_borrower = wb.create_sheet("Borrower Distress")
    borrower_df = (
        df[df["layer"] == "borrower"]
        .sort_values("stress_score", ascending=False)
        .reset_index(drop=True)
    )
    borrower_topic_cols = [
        t["topic_name"]
        for t in TOPICS
        if "borrower" in t["applies_to"] and t["polarity"] == "negative"
    ]
    borrower_cols = ["entity_name", "ticker", "stress_score"] + [
        c for c in borrower_topic_cols if c in borrower_df.columns
    ]
    _write_ranked_sheet(ws_borrower, borrower_df, borrower_cols, score_col="stress_score")

    # ── Sheet 3: Bank Contagion (optional) ───────────────────────────────────
    if INCLUDE_BANKS_LAYER:
        ws_bank = wb.create_sheet("Bank Contagion")
        bank_df = df[df["layer"] == "bank"].copy()
        _gain = (
            bank_df["bank_market_share_gain"].fillna(0)
            if "bank_market_share_gain" in bank_df.columns
            else 0
        )
        _pullback = (
            bank_df["bank_credit_pullback"].fillna(0)
            if "bank_credit_pullback" in bank_df.columns
            else 0
        )
        bank_df["net_position_score"] = _gain - _pullback
        bank_df = bank_df.sort_values("net_position_score", ascending=False).reset_index(
            drop=True
        )
        bank_topic_cols = [t["topic_name"] for t in TOPICS if "bank" in t["applies_to"]]
        bank_cols = ["entity_name", "ticker", "net_position_score"] + [
            c for c in bank_topic_cols if c in bank_df.columns
        ]
        _write_ranked_sheet(ws_bank, bank_df, bank_cols, score_col="net_position_score")

    # ── Sheets 4–6: Raw Signal Matrix (split by layer) ─────────────────────────
    for layer_key, layer_label in [("lender", "Lenders"), ("borrower", "Borrowers"), ("bank", "Banks")]:
        layer_topics = [str(t["topic_name"]) for t in TOPICS if layer_key in t["applies_to"]]
        matrix_cols = ["entity_name", "ticker"] + [c for c in layer_topics if c in df.columns]
        layer_df = df[df["layer"] == layer_key].copy()
        if layer_df.empty:
            continue
        ws_matrix = wb.create_sheet(f"Raw Signal Matrix - {layer_label}")
        _write_matrix_sheet(ws_matrix, layer_df, matrix_cols, use_topic_display_names=True, layer_key=layer_key)

    # ── Methodology ─────────────────────────────────────────────────────────
    ws_method = wb.create_sheet("Methodology")
    _write_methodology_sheet(ws_method)

    EXCEL_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(EXCEL_OUTPUT))
    logger.info("Excel report saved to %s", EXCEL_OUTPUT)
    return EXCEL_OUTPUT


def _write_ranked_sheet(
    ws: Any,
    df: pd.DataFrame,
    columns: list[str],
    score_col: str,
) -> None:
    available_cols = [c for c in columns if c in df.columns]
    ws.append(["Rank"] + [c.replace("_", " ").title() for c in available_cols])
    _style_header(ws)

    for rank, (_, row) in enumerate(df.iterrows(), 1):
        values = [rank] + [row.get(c) for c in available_cols]
        ws.append(values)

    n_rows = len(df)
    n_cols = len(available_cols) + 1
    _style_body(ws, n_rows, n_cols)

    score_col_idx = available_cols.index(score_col) + 2 if score_col in available_cols else None
    if score_col_idx and n_rows > 1:
        col_letter = get_column_letter(score_col_idx)
        ws.conditional_formatting.add(
            f"{col_letter}2:{col_letter}{n_rows + 1}",
            ColorScaleRule(
                start_type="min", start_color="FF6B6B",
                mid_type="percentile", mid_value=50, mid_color="FFD93D",
                end_type="max", end_color="6BCB77",
            ),
        )

    _auto_width(ws)


def _write_matrix_sheet(
    ws: Any,
    df: pd.DataFrame,
    columns: list[str],
    use_topic_display_names: bool = False,
    layer_key: str = "",
) -> None:
    available_cols = [c for c in columns if c in df.columns]
    headers = (
        [_topic_display_name(c) for c in available_cols]
        if use_topic_display_names
        else [c.replace("_", " ").title() for c in available_cols]
    )
    ws.append(headers)
    _style_header(ws)

    polarity_map = {str(t["topic_name"]): str(t["polarity"]) for t in TOPICS}

    # Colour headers to match UI: red header font for negative, teal for positive
    neg_header_font = Font(bold=True, color="C62828", size=11)
    pos_header_font = Font(bold=True, color="00796B", size=11)
    for col_idx, col_name in enumerate(available_cols, 1):
        pol = polarity_map.get(col_name, "")
        if pol == "negative":
            ws.cell(row=1, column=col_idx).font = neg_header_font
        elif pol == "positive":
            ws.cell(row=1, column=col_idx).font = pos_header_font

    for _, row in df.iterrows():
        ws.append([row.get(c) for c in available_cols])

    n_rows = len(df)
    n_cols = len(available_cols)
    _style_body(ws, n_rows, n_cols)

    topic_start_col = 3
    if n_rows > 0 and n_cols >= topic_start_col:
        for col_idx in range(topic_start_col, n_cols + 1):
            col_letter = get_column_letter(col_idx)
            col_name = available_cols[col_idx - 1]
            pol = polarity_map.get(col_name, "")
            if pol == "negative":
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{n_rows + 1}",
                    ColorScaleRule(
                        start_type="min", start_color="FFFFFF",
                        end_type="max", end_color="C62828",
                    ),
                )
            else:
                ws.conditional_formatting.add(
                    f"{col_letter}2:{col_letter}{n_rows + 1}",
                    ColorScaleRule(
                        start_type="min", start_color="FFFFFF",
                        end_type="max", end_color="00796B",
                    ),
                )

    _auto_width(ws)


def _write_methodology_sheet(ws: Any) -> None:
    methodology_text = [
        ["Private Credit Stress Analyzer — Methodology"],
        [""],
        ["Data Source:"],
        ["  Bigdata API semantic search across news, filings, and transcripts."],
        [""],
        ["Entity Layers:"],
        ["  1. Lenders — BDCs, private credit funds, alternative asset managers"],
        ["  2. Borrowers — PE-backed companies with leveraged loan exposure"],
        ["  3. Banks — Back-leverage providers to private credit funds"],
        [""],
        ["Search Topics:"],
        ["  Each entity is searched against topic-specific query texts."],
        ["  Topics have polarity: 'positive' (strength) or 'negative' (stress)."],
        ["  {company} placeholder is replaced with entity name at runtime."],
        ["  Retrieval: each query uses a Bigdata sentiment filter aligned to that topic's polarity"],
        ["  (positive topics favor positive-toned documents; negative topics favor negative-toned)."],
        [""],
        ["Scoring Formula:"],
        ["  terms_power_score = positive_count / (positive_count + negative_count + 1) × 100"],
        ["  stress_score = 100 - terms_power_score"],
        [""],
        ["Layer-Specific Ranking:"],
        ["  Lenders: ranked by terms_power_score (high = strong)"],
        ["  Borrowers: ranked by stress_score (high = distressed)"],
        ["  Banks: ranked by net_position = market_share_gain − credit_pullback"],
        [""],
        ["Limitations:"],
        ["  Scores are mention counts after entity-in-text filtering, not chunk sentiment scores."],
        ["  Results depend on Bigdata index coverage and recency."],
    ]
    for row in methodology_text:
        ws.append(row)

    ws.column_dimensions["A"].width = 80
    for row_cells in ws.iter_rows():
        for cell in row_cells:
            cell.font = Font(color="000000", size=11)
            cell.fill = BODY_FILL
    ws["A1"].font = Font(bold=True, color="000000", size=14)


def _load_audit_docs(
    layer: str,
    entities: list[dict[str, str | None]] | None = None,
    audit_dir: Path | None = None,
) -> list[dict[str, Any]]:
    """Load documents from scoring output (``.cache/scoring_audit/``).

    Rows are exactly those persisted when ``compute_scores`` ran—single source of truth
    with heatmap counts. Run the scorer before generating HTML.
    """
    from config.entities import ALL_ENTITIES
    from src.utils import sanitize_filename

    layer_entities = [e for e in (entities or ALL_ENTITIES) if e["layer"] == layer]
    layer_topics = [t for t in TOPICS if layer in t["applies_to"]]
    _audit_dir = audit_dir if audit_dir is not None else SCORING_AUDIT_DIR
    docs: list[dict[str, Any]] = []

    for entity in layer_entities:
        entity_slug = sanitize_filename(str(entity["name"]))
        for topic in layer_topics:
            topic_slug = sanitize_filename(str(topic["topic_name"]))
            path = _audit_dir / f"{entity_slug}_{topic_slug}.json"
            if not path.exists():
                continue
            payload = json.loads(path.read_text())
            topic_display = str(topic.get("short_label", topic["topic_name"])).strip()
            polarity = payload.get("topic_polarity", topic["polarity"])
            ename = str(payload.get("entity_name", entity["name"]))
            for result in payload.get("results", []):
                content = result.get("content", "")
                docs.append({
                    "entity": ename,
                    "topic": topic_display,
                    "polarity": polarity,
                    "headline": result.get("headline", ""),
                    "snippet": content[:250] + ("..." if len(content) > 250 else ""),
                    "timestamp": (result.get("timestamp") or "")[:10],
                    "url": result.get("url", ""),
                })
    return docs


def _prepare_layer_data(
    df: pd.DataFrame,
    layer: str,
    entities: list[dict[str, str | None]] | None = None,
    audit_dir: Path | None = None,
) -> dict[str, Any]:
    """Prepare all chart/heatmap/theme data for a single layer."""
    layer_df = df[df["layer"] == layer].copy()
    layer_topics = [t for t in TOPICS if layer in t["applies_to"]]
    topic_order_index = {str(t["topic_name"]): i for i, t in enumerate(TOPICS)}
    layer_topics_sorted = sorted(
        layer_topics,
        key=lambda t: (
            0 if str(t["polarity"]) == "positive" else 1,
            topic_order_index[str(t["topic_name"])],
        ),
    )

    if layer == "lender":
        score_col = "terms_power_score"
        layer_df = layer_df.sort_values(score_col, ascending=True)
    elif layer == "borrower":
        score_col = "stress_score"
        layer_df = layer_df.sort_values(score_col, ascending=False)
    else:
        gain = layer_df["bank_market_share_gain"].fillna(0) if "bank_market_share_gain" in layer_df.columns else 0
        pullback = layer_df["bank_credit_pullback"].fillna(0) if "bank_credit_pullback" in layer_df.columns else 0
        layer_df["net_position"] = gain - pullback
        score_col = "net_position"
        layer_df = layer_df.sort_values(score_col, ascending=True)

    labels = layer_df["entity_name"].tolist()
    scores = layer_df[score_col].tolist()

    heatmap_rows: list[list[int]] = []
    topic_names_sorted = [str(t["topic_name"]) for t in layer_topics_sorted]
    available_topics = [t for t in topic_names_sorted if t in layer_df.columns]
    polarity_by_topic = {str(t["topic_name"]): str(t["polarity"]) for t in layer_topics}
    heatmap_topic_polarities = [polarity_by_topic[t] for t in available_topics]
    for _, row in layer_df.iterrows():
        heatmap_rows.append([
            int(v) if pd.notna(v) else 0
            for t in available_topics
            for v in [row.get(t, 0)]
        ])

    theme_topics: list[dict[str, Any]] = []
    for topic in layer_topics_sorted:
        tname = str(topic["topic_name"])
        total = int(layer_df[tname].fillna(0).sum()) if tname in layer_df.columns else 0
        theme_topics.append({
            "short_label": str(topic["short_label"]),
            "description": _build_topic_description(topic),
            "polarity": topic["polarity"],
            "query": str(topic["topic_text"]),
            "count": total,
        })
    positive_themes = [x for x in theme_topics if x["polarity"] == "positive"]
    negative_themes = [x for x in theme_topics if x["polarity"] == "negative"]
    positive_themes.sort(key=lambda x: x["count"], reverse=True)
    negative_themes.sort(key=lambda x: x["count"], reverse=True)
    theme_topics = positive_themes + negative_themes

    audit_docs = _load_audit_docs(layer, entities=entities, audit_dir=audit_dir)

    out: dict[str, Any] = {
        "labels": labels,
        "scores": scores,
        "score_col": score_col,
        "heatmap_entities": labels,
        "heatmap_topics": [str(TOPIC_BY_NAME[t]["short_label"]) for t in available_topics],
        "heatmap_topic_polarities": heatmap_topic_polarities,
        "heatmap_data": heatmap_rows,
        "theme_topics": theme_topics,
        "audit_docs": audit_docs,
        "entity_count": len(labels),
        "topic_count": len(layer_topics_sorted),
    }

    # Radar: build from negative topics so radar and heatmap share the same raw counts
    if layer in ("borrower", "lender"):
        radar_colors = ["#FF6B6B", "#FFD93D", "#4ECDC4", "#45B7D1", "#96CEB4", "#FFEAA7"]
        negative_topics = [
            str(t["topic_name"])
            for t in layer_topics_sorted
            if t["polarity"] == "negative" and str(t["topic_name"]) in available_topics
        ]
        radar_datasets = []
        for i, topic in enumerate(negative_topics):
            idx = available_topics.index(topic)
            radar_datasets.append({
                "label": str(TOPIC_BY_NAME[topic]["short_label"]),
                "data": [heatmap_rows[row_idx][idx] for row_idx in range(len(heatmap_rows))],
                "borderColor": radar_colors[i % len(radar_colors)],
                "backgroundColor": radar_colors[i % len(radar_colors)] + "33",
            })
        out["radar_datasets"] = radar_datasets

    return out


def generate_html_dashboard(df: pd.DataFrame) -> Path:
    """Generate a standalone HTML dashboard with Chart.js visualizations."""
    lender_data = _prepare_layer_data(df, "lender")
    borrower_data = _prepare_layer_data(df, "borrower")
    bank_data = _prepare_layer_data(df, "bank")

    # Radar data (JSON-encoded for JS injection)
    lender_radar_labels = json.dumps(lender_data["labels"])
    lender_radar_datasets = json.dumps(lender_data.get("radar_datasets", []))
    borrower_radar_labels = json.dumps(borrower_data["labels"])
    borrower_radar_datasets = json.dumps(borrower_data.get("radar_datasets", []))

    html = _build_html(
        lender_data=lender_data,
        borrower_data=borrower_data,
        bank_data=bank_data,
        lender_radar_labels=lender_radar_labels,
        lender_radar_datasets=lender_radar_datasets,
        borrower_radar_labels=borrower_radar_labels,
        borrower_radar_datasets=borrower_radar_datasets,
        include_banks=INCLUDE_BANKS_LAYER,
    )

    HTML_INDEX.parent.mkdir(parents=True, exist_ok=True)
    HTML_INDEX.write_text(html)
    (DIST_DIR / ".nojekyll").write_text("", encoding="utf-8")
    logger.info("HTML dashboard saved to %s", HTML_INDEX)
    return HTML_INDEX


def _themes_html(topics: list[dict[str, Any]]) -> str:
    """Build the Key Themes section with polarity, count, and inline description."""
    items = ""
    for t in topics:
        pol_class = "positive" if t["polarity"] == "positive" else "negative"
        pol_icon = "+" if t["polarity"] == "positive" else "&minus;"
        query_escaped = str(t["query"]).replace("{company}", "<em>{company}</em>")
        label = html.escape(str(t["short_label"]))
        desc = html.escape(str(t["description"]), quote=False)
        items += (
            f'<div class="theme-card {pol_class}">'
            f'<div class="theme-header">'
            f'<span class="theme-name">{label} '
            f'<span class="theme-desc">({desc})</span></span>'
            f'<span class="pol-badge {pol_class[:3]}">{pol_icon} {t["polarity"].title()}</span>'
            f'<span class="theme-count">{t["count"]}</span>'
            f'</div>'
            f'<div class="theme-query"><span class="theme-query-label">Search Query:</span> {query_escaped}</div>'
            f'</div>'
        )
    return items


def _build_html(
    lender_data: dict[str, Any],
    borrower_data: dict[str, Any],
    bank_data: dict[str, Any],
    lender_radar_labels: str,
    lender_radar_datasets: str,
    borrower_radar_labels: str,
    borrower_radar_datasets: str,
    *,
    include_banks: bool,
) -> str:
    favicon_svg = "data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 159.7 159.7'%3E%3Cpath fill='%234CA7F9' d='M38.11,0h83.48c21.03,0,38.11,17.08,38.11,38.11v83.48c0,21.03-17.08,38.11-38.11,38.11H38.11c-21.03,0-38.11-17.08-38.11-38.11V38.11C0,17.08,17.08,0,38.11,0Z'/%3E%3Cpath fill='%23FFFDF5' d='M105.69,137.06c-8.4,0-16.35-3.27-22.4-9.21-6.07-5.96-9.41-13.84-9.41-22.18v-51.63c0-11.38-8.72-20.3-19.87-20.3-5.4,0-10.44,2.12-14.21,5.96-3.74,3.82-5.81,8.91-5.81,14.34s2.06,10.52,5.81,14.34c3.76,3.84,8.81,5.96,14.21,5.96h13.36v11.09h-15.49c-8.26,0-15.83-3.26-21.32-9.19-5.4-5.83-8.37-13.71-8.37-22.2s3.34-16.22,9.41-22.18c6.05-5.94,14-9.21,22.4-9.21s16.36,3.27,22.4,9.21c6.07,5.96,9.41,13.84,9.41,22.18v51.63c0,11.38,8.72,20.3,19.86,20.3,5.4,0,10.44-2.12,14.21-5.96,3.74-3.82,5.81-8.91,5.81-14.34s-2.06-10.52-5.81-14.34c-3.76-3.84-8.81-5.96-14.21-5.96h-13.26v-11.09h15.4c8.26,0,15.83,3.26,21.32,9.19,5.4,5.82,8.37,13.71,8.37,22.2s-3.34,16.22-9.41,22.18c-6.05,5.94-14,9.21-22.4,9.21Z'/%3E%3C/svg%3E"

    # Per-layer descriptions for the 4 stat tiles (Entities, Topics, Highest, Lowest)
    _stat_descriptions = {
        "lender": (
            "BDCs and private credit managers in the universe.",
            "Strength and stress topics used for semantic search.",
            "Best terms power score across lenders (0–100).",
            "Lowest terms power; most stress signals.",
        ),
        "borrower": (
            "PE-backed companies with leveraged loan exposure.",
            "Resilience and distress topics per borrower.",
            "Most distressed borrower (highest stress score).",
            "Least distressed in the cohort.",
        ),
        "bank": (
            "Banks providing back-leverage to private credit.",
            "Contagion and market-share signal topics.",
            "Best net position (gain vs pullback).",
            "Worst net position; most pullback signal.",
        ),
    }

    _chart_descriptions = {
        "lender_score": (
            "Terms Power Score (0–100) = positive_mentions / (positive + negative + 1) × 100. "
            "Higher = stronger lender signal (spread power, fundraise resilience, NAV stability, covenant tightening "
            "vs redemption pressure, markdowns, waivers, PIK stress). Bars ranked by score descending."
        ),
        "lender_radar": (
            "Each line is one stress topic (redemption pressure, NAV markdown, covenant waiver, PIK stress, fee compression). "
            "Value at each axis = mention count for that lender × topic. Same raw counts as the Signal Heatmap; only negative topics shown."
        ),
        "borrower_radar": (
            "Each line is one distress topic (AI disruption, maturity wall, default risk, customer churn). "
            "Value at each axis = mention count for that borrower × topic. Same raw counts as the Signal Heatmap; only negative topics shown."
        ),
        "borrower_stress": (
            "Stress score (0–100) = 100 − terms_power_score, where terms_power = positive / (positive + negative + 1) × 100. "
            "High score = more distress signal relative to resilience (revenue growth, refinancing). Ranked by stress descending."
        ),
        "bank": (
            "Net position = market_share_gain_mentions − credit_pullback_mentions (and related negative topics). "
            "Positive = bank gaining share or conservative exposure; negative = pullback or contagion risk. Bars ranked by net position descending."
        ),
    }
    _heatmap_descriptions = {
        "lender": (
            "Each cell = number of returned documents where the lender’s name appears in the headline or content for that topic query "
            "(not the raw semantic-search list size). Rows = entities, columns = topics; these counts feed the Terms Power Score. "
            "Audit lists the same rows saved under .cache/scoring_audit/ when scores were computed."
        ),
        "borrower": (
            "Each cell = number of returned documents where the borrower’s name appears in the headline or content for that topic—same rule as scoring. "
            "Positive columns (revenue growth, refinancing, etc.) and negative columns (AI disruption, maturity wall, default risk, customer churn) "
            "combine into the stress score as a ratio, not a raw sum. Audit reads the scoring snapshot (.cache/scoring_audit/), not raw search."
        ),
        "bank": (
            "Each cell = number of returned documents where the bank’s name appears in the headline or content for that topic. "
            "Net position uses the delta between positive and negative topic counts. Audit uses the same scoring snapshot files."
        ),
    }

    lender_tab_intros: dict[str, str] = {
        "layer": (
            "Who’s showing lender-strength narratives versus stress (redemptions, markdowns, waivers) in recent news and filings."
        ),
        "chart": (
            "The radar plots intensity on stress topics only; the horizontal bars are Terms Power Score—higher means more "
            "strength-themed mentions relative to stress for that lender."
        ),
        "heatmap": (
            "Each cell counts snippets that matched the theme query and mention the lender in headline or body. "
            "Teal-headed columns are strength; coral-headed columns are stress—the same counts roll up into the score."
        ),
        "themes": (
            "Every theme is a fixed semantic search. Strength topics are listed first, then stress (green block above red). "
            "Each card shows its definition and scoring direction in parentheses."
        ),
        "audit": (
            "Supporting headlines and excerpts from the scoring run. Filter by entity or theme to verify what drove a cell."
        ),
        "method": (
            "End-to-end methodology: search, entity filter, polarity aggregation, and the Terms Power formula for this layer."
        ),
    }
    borrower_tab_intros: dict[str, str] = {
        "layer": (
            "Which PE-backed names skew toward resilience stories versus distress (maturity wall, default risk, churn, AI headwinds)."
        ),
        "chart": (
            "The radar highlights distress-topic volume by borrower; the stress bars combine positive and negative themes—higher means more distress signal."
        ),
        "heatmap": (
            "Per-borrower, per-theme mention counts after the entity must appear in the returned text. Resilience columns first, then distress."
        ),
        "themes": (
            "Queries used for borrowers: resilience themes on top, distress below. Each card shows its definition and scoring direction."
        ),
        "audit": (
            "Documents that passed the same filters used in scoring—use this tab to sanity-check a heatmap cell."
        ),
        "method": (
            "How borrower stress_score is built from positive versus negative theme counts and Bigdata retrieval settings."
        ),
    }
    bank_tab_intros: dict[str, str] = {
        "layer": (
            "Banks as the back-leverage layer: net position contrasts constructive share-gain narratives against pullback and contagion mentions."
        ),
        "chart": (
            "Horizontal bars show net position (positive-topic mentions minus pullback-style mentions)—compare banks side by side."
        ),
        "heatmap": (
            "Raw mention counts per bank and theme; positive themes first, then stress. Feeds the same net position idea as the bar chart."
        ),
        "themes": (
            "Themes specific to bank/private-credit linkage. Green block first, red below. Each card shows its definition."
        ),
        "audit": (
            "Source rows for bank × theme counts; filter to trace a headline behind a heatmap value."
        ),
        "method": (
            "Methodology for bank-layer searches and net position scoring."
        ),
    }

    _cov = (
        f"Search date filter: <strong>{SEARCH_DATE_LABEL}</strong>. Sources: news, filings, transcripts indexed by "
        f"<a href='https://bigdata.com' target='_blank' rel='noopener' style='color:#4CA7F9;text-decoration:none;'>Bigdata.com</a>."
    )

    overview_inner = f"""
    <div class="overview-hero">
      <h2>What it does</h2>
      <p class="overview-lead">The analyzer is a thematic research tool built on <a href="https://bigdata.com" target="_blank" rel="noopener">Bigdata.com</a>. It:</p>
      <ul class="overview-list">
        <li>Searches news, filings, and transcripts for a configurable set of lenders and borrowers against predefined thematic queries (e.g. spread power, redemption pressure, maturity wall risk).</li>
        <li>Scores each entity by comparing how many on-topic, entity-relevant mentions fall into positive themes versus negative themes.</li>
      </ul>
    </div>

    <div class="overview-section">
        
      <h3>Configurable</h3>
      <ul class="overview-list">
        <li> (Required re-run) </li>
      </ul>
      <ul class="overview-config-list">
        <li><span class="config-label">Entities</span> Lender and borrower names are configurable.</li>
        <li><span class="config-label">Themes</span> Search topics and their queries are configurable; You can define the risk and strength signals that matter to your investment process.</li>
        <li><span class="config-label">Scoring</span> The scoring formula can be replaced with your own weighting, composite, or rules-based model.</li>
        <li><span class="config-label">Date range</span> The search window is configurable; narrow to the last 30 days or widen to multi-year history.</li>
      </ul>
    </div>

    <div class="overview-section overview-disclaimer">
      <h3>Important</h3>
      <p>This is a technical showcase of <a href="https://bigdata.com" target="_blank" rel="noopener">Bigdata.com</a> capabilities, not investment advice, a credit rating, or a trading signal. Scores are derived from mention counts after entity-in-text filtering and depend on the configured entities, topics, date range, and how content is indexed. Independent verification is recommended before any investment decision.</p>
    </div>

    <h2 class="overview-scores-heading">How to read scores</h2>
    <p class="overview-scores-lead">Counts are on-topic snippets where the entity name appears in the returned text, not raw unfiltered search volume.</p>
    <div class="overview-scores-grid">
      <div class="overview-card">
        <h3>Lender — Terms Power Score</h3>
        <p><code>terms_power_score = positive_count / (positive_count + negative_count + 1) × 100</code></p>
        <p>Positive themes include spread power, fundraise resilience, NAV stability, and covenant tightening. Negative themes include redemption pressure, NAV markdown, covenant waivers, PIK stress, and fee compression.</p>
        <p><strong>Higher score = stronger lender positioning.</strong> A score near 100 means almost all entity-relevant mentions align with strength themes; a score near 0 means stress themes dominate.</p>
      </div>
      <div class="overview-card">
        <h3>Borrower — Stress Score</h3>
        <p><code>stress_score = 100 − terms_power_score</code></p>
        <p>Positive (resilience) themes include revenue growth, refinancing success, margin expansion, and AI adoption. Negative (distress) themes include AI disruption risk, maturity wall, default/restructuring, and customer churn.</p>
        <p><strong>Higher score = greater distress signal.</strong> A score near 100 means distress themes heavily outweigh resilience mentions; a low score means the borrower is showing more positive coverage.</p>
      </div>
    </div>
    <p class="overview-footer">{_cov}</p>
    """

    overview_page = f"""
    <div class="layer-page active" id="page-overview">
      <div class="overview-wrap">{overview_inner}</div>
    </div>"""

    def _tab_intro_line(text: str) -> str:
        return f'<p class="tab-intro">{text}</p>' if text else ""

    def _layer_page(
        layer: str,
        data: dict[str, Any],
        chart_label: str,
        tab1_name: str,
        method_html: str,
        chart_panel_html: str | None = None,
        heatmap_desc: str = "",
        *,
        tab_intros: dict[str, str],
        page_active_class: str = "",
    ) -> str:
        scores = data["scores"]
        hi = max(scores) if scores else 0
        lo = min(scores) if scores else 0
        hi_label = "Highest Score" if layer == "lender" else "Highest Stress" if layer == "borrower" else "Highest Net Pos."
        lo_label = "Lowest Score" if layer == "lender" else "Lowest Stress" if layer == "borrower" else "Lowest Net Pos."
        descs = _stat_descriptions[layer]
        active = page_active_class
        if chart_panel_html is None:
            chart_panel_html = f'<div class="card"><h3>{chart_label}</h3><p class="card-desc">{_chart_descriptions.get(layer, "")}</p><canvas id="{layer}Chart"></canvas></div>'
        heatmap_block = f'<p class="card-desc">{heatmap_desc}</p>' if heatmap_desc else ''
        ti = tab_intros
        layer_blurb = (
            f'<p class="layer-blurb">{ti["layer"]}</p>' if ti.get("layer") else ""
        )
        return f"""
    <div class="layer-page{active}" id="page-{layer}">
      {layer_blurb}
      <div class="stats">
        <div class="stat-card"><div class="stat-val">{data["entity_count"]}</div><div class="stat-label">Entities Tracked</div><div class="stat-desc">{descs[0]}</div></div>
        <div class="stat-card"><div class="stat-val">{data["topic_count"]}</div><div class="stat-label">Signal Topics</div><div class="stat-desc">{descs[1]}</div></div>
        <div class="stat-card"><div class="stat-val">{hi:.1f}</div><div class="stat-label">{hi_label}</div><div class="stat-desc">{descs[2]}</div></div>
        <div class="stat-card"><div class="stat-val">{lo:.1f}</div><div class="stat-label">{lo_label}</div><div class="stat-desc">{descs[3]}</div></div>
      </div>
      <div class="tabs" data-tabgroup="{layer}">
        <div class="tab active" data-tab="chart" onclick="switchTab('{layer}','chart')">{tab1_name}</div>
        <div class="tab" data-tab="heatmap" onclick="switchTab('{layer}','heatmap')">Signal Heatmap</div>
        <div class="tab" data-tab="themes" onclick="switchTab('{layer}','themes')">Key Themes</div>
        <div class="tab" data-tab="audit" onclick="switchTab('{layer}','audit')">Audit</div>
        <div class="tab" data-tab="method" onclick="switchTab('{layer}','method')">Methodology</div>
      </div>
      <div class="tab-panel active" id="{layer}-chart">{_tab_intro_line(ti.get("chart", ""))}{chart_panel_html}</div>
      <div class="tab-panel" id="{layer}-heatmap">{_tab_intro_line(ti.get("heatmap", ""))}<div class="card"><h3>Signal Matrix</h3>{heatmap_block}<div class="heatmap-wrap" id="{layer}Heatmap"></div></div></div>
      <div class="tab-panel" id="{layer}-themes">{_tab_intro_line(ti.get("themes", ""))}<div class="card"><h3>Signal Topics &amp; Queries</h3><div class="themes-list">{_themes_html(data["theme_topics"])}</div></div></div>
      <div class="tab-panel" id="{layer}-audit">{_tab_intro_line(ti.get("audit", ""))}<div class="card"><h3>Document Audit</h3><div class="audit-filters" id="{layer}AuditFilters"></div><div class="audit-count" id="{layer}AuditCount"></div><div class="audit-table-wrap" id="{layer}AuditTable"></div></div></div>
      <div class="tab-panel" id="{layer}-method">{_tab_intro_line(ti.get("method", ""))}<div class="method-block">{method_html}</div></div>
    </div>"""

    def _method_page(title: str, subtitle: str, why: str, what: str, formula_code: str, formula_desc: str, steps: list[tuple[str, str, str, str]], coverage: str) -> str:
        steps_html = ""
        colors = ["blue", "green", "amber", "red", "blue", "green"]
        for i, (step_title, step_desc, step_detail, _) in enumerate(steps):
            detail_html = f'<div class="step-detail">{step_detail}</div>' if step_detail else ''
            steps_html += f'<div class="method-step"><div class="method-step-num {colors[i % len(colors)]}">{i+1}</div><h4>{step_title}</h4><p>{step_desc}</p>{detail_html}</div>'
        return (
            f'<div class="method-hero"><h2>{title}</h2><p>{subtitle}</p></div>'
            f'<div class="method-cards">'
            f'<div class="method-card"><h3>&#10024; Why It Matters</h3><p>{why}</p></div>'
            f'<div class="method-card"><h3>&#9889; What It Does</h3><p>{what}</p></div>'
            f'</div>'
            f'<div class="method-hero" style="margin-bottom:1.2rem"><h2>The Analysis Process</h2></div>'
            f'<div class="method-steps">{steps_html}</div>'
            f'<div class="method-cards" style="margin-top:1.5rem;grid-template-columns:1fr">'
            f'<div class="method-card"><h3>&#128290; Scoring Formula</h3>'
            f'<p><code>{formula_code}</code></p><p style="margin-top:0.4rem">{formula_desc}</p></div></div>'
            f'<div class="method-cards" style="margin-top:0.5rem;grid-template-columns:1fr">'
            f'<div class="method-card"><h3>&#128197; Coverage</h3><p>{coverage}</p></div></div>'
        )

    lender_method = _method_page(
        "How Lender Scoring Works",
        "The system combines <strong>hybrid semantic search</strong>, <strong>topic taxonomies</strong>, and <strong>polarity-based scoring</strong> to measure lender health.",
        "Identifying which lenders hold strong negotiating positions versus those under redemption, markdown, or PIK stress is critical for understanding systemic private credit risk.",
        "Each lender is searched against strength topics (spread power, fundraise resilience, NAV stability, covenant tightening) and stress topics (redemption pressure, NAV markdown, covenant waivers, PIK stress, fee compression). Mention counts feed a Terms Power Score.",
        "terms_power_score = positive / (positive + negative + 1) &times; 100",
        "High score = strong lender position. Low score = elevated stress signals relative to strength.",
        [
            ("Topic Search", "Each lender is searched against every signal topic using Bigdata semantic search. A sentiment filter matches each topic's polarity—positive topics retrieve predominantly positive-toned documents; stress topics retrieve predominantly negative-toned documents (same idea as API sentiment filters).", "Example: \"Blackstone\" &times; \"redemption pressure\" &rarr; thematic document retrieval with negative sentiment alignment", ""),
            ("Entity Filtering", "Results are filtered to only include documents where the entity name appears in headline or content.", "Ensures the returned text gives a true relevance signal that differentiates entities", ""),
            ("Polarity Aggregation", "Mention counts are split by polarity (positive / negative) and summed per entity.", "", ""),
            ("Scoring &amp; Ranking", "Terms Power Score is computed and lenders are ranked descending. Radar chart shows negative-topic signal intensity.", "", ""),
        ],
        _cov,
    )

    borrower_method = _method_page(
        "How Borrower Distress Scoring Works",
        "The system combines <strong>hybrid semantic search</strong>, <strong>risk factor taxonomies</strong>, and <strong>structured validation</strong> to transform unstructured data into actionable distress intelligence.",
        "Understanding which PE-backed borrowers face AI disruption, maturity walls, default risk, or customer churn is critical for private credit portfolio monitoring.",
        "Each borrower is searched against resilience topics (revenue growth, refinancing success) and distress topics (AI disruption, maturity wall, default risk, customer churn). The ratio drives a Stress Score.",
        "stress_score = 100 &minus; terms_power_score",
        "High stress = more distress signal relative to resilience. Low stress = borrower showing strength signals.",
        [
            ("Topic Search", "Each borrower is searched against every signal topic using Bigdata semantic search. A sentiment filter matches each topic's polarity—resilience topics favor positive-toned documents; distress topics favor negative-toned documents (same idea as API sentiment filters).", "Example: \"Cision\" &times; \"AI disruption risk\" &rarr; thematic document retrieval with negative sentiment alignment", ""),
            ("Entity Filtering", "Results are filtered to only include documents that explicitly mention the borrower.", "Filters out content that doesn't explicitly link companies to risk factors", ""),
            ("Polarity Aggregation", "Mention counts are split by polarity and summed. Radar chart uses only negative topics.", "", ""),
            ("Scoring &amp; Ranking", "Stress score is computed and borrowers are ranked descending. High score = most distressed.", "", ""),
        ],
        _cov + " Focus on PE-backed leveraged companies with private credit exposure.",
    )

    bank_method = _method_page(
        "How Bank Contagion Scoring Works",
        "The system measures bank <strong>back-leverage exposure</strong> to private credit using <strong>net position scoring</strong>.",
        "Banks provide leverage to lenders. Understanding which banks are pulling back versus gaining share reveals contagion risk pathways.",
        "Each bank is searched against contagion topics (credit pullback, margin calls, contagion risk) and positive topics (market share gain). Net position = gain minus pullback.",
        "net_position = market_share_gain &minus; credit_pullback",
        "Positive = bank gaining share or conservative. Negative = pulling back or contagion risk.",
        [
            ("Topic Search", "Each bank is searched against back-leverage, contagion, and market share topics using Bigdata semantic search, with sentiment aligned to each topic's polarity (positive vs negative).", "Searches across multiple document types with configurable date ranges", ""),
            ("Entity Filtering", "Results are filtered to only include documents mentioning the bank in headline or content.", "", ""),
            ("Net Position", "Net position = delta between positive and negative topic mentions.", "", ""),
            ("Ranking", "Banks ranked by net position (descending). Positive = gaining share.", "", ""),
        ],
        _cov + " Focus on banks providing back-leverage to private credit funds.",
    )

    lender_chart_panel = (
        '<div class="card"><h3>Stress radar</h3><p class="card-desc">'
        + _chart_descriptions["lender_radar"]
        + '</p><canvas id="lenderRadarChart"></canvas></div>'
        + '<div class="card"><h3>Terms Power Score by lender</h3><p class="card-desc">'
        + _chart_descriptions["lender_score"]
        + '</p><div class="lender-score-chart-wrap"><canvas id="lenderChart"></canvas></div></div>'
    )
    lender_page = _layer_page(
        "lender",
        lender_data,
        "Lender Terms Power Score",
        "Score Analysis",
        lender_method,
        chart_panel_html=lender_chart_panel,
        heatmap_desc=_heatmap_descriptions["lender"],
        tab_intros=lender_tab_intros,
        page_active_class="",
    )
    borrower_chart_panel = (
        '<div class="card"><h3>Distress radar</h3><p class="card-desc">'
        + _chart_descriptions["borrower_radar"]
        + '</p><canvas id="borrowerChart"></canvas></div>'
        + '<div class="card"><h3>Stress score by company</h3><p class="card-desc">'
        + _chart_descriptions["borrower_stress"]
        + '</p><div class="borrower-score-chart-wrap"><canvas id="borrowerScoreChart"></canvas></div></div>'
    )
    borrower_page = _layer_page(
        "borrower",
        borrower_data,
        "Distress",
        "Distress",
        borrower_method,
        chart_panel_html=borrower_chart_panel,
        heatmap_desc=_heatmap_descriptions["borrower"],
        tab_intros=borrower_tab_intros,
        page_active_class="",
    )

    le_ct = lender_data["entity_count"]
    le_tp = lender_data["topic_count"]
    bo_ct = borrower_data["entity_count"]
    bo_tp = borrower_data["topic_count"]
    if include_banks:
        bank_page_html = _layer_page(
            "bank",
            bank_data,
            "Bank Net Position Score",
            "Contagion Score",
            bank_method,
            heatmap_desc=_heatmap_descriptions["bank"],
            tab_intros=bank_tab_intros,
            page_active_class="",
        )
        bank_nav_html = """
    <div class="nav-item" data-layer="bank" onclick="switchLayer('bank')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 21h18M3 10h18M5 6l7-3 7 3M4 10v11M20 10v11M8 14v3M12 14v3M16 14v3"/></svg>
      Banks
    </div>"""
        page_titles_js = (
            "{ overview:'Private Credit Stress Analyzer', lender:'Lender Terms Power Analysis', "
            "borrower:'Borrower Distress Analysis', bank:'Bank Contagion Analysis' }"
        )
        ba_ct = bank_data["entity_count"]
        ba_tp = bank_data["topic_count"]
        page_badges_js = (
            f"{{ overview:'Overview &middot; thematic demo', lender:'{le_ct} entities &middot; {le_tp} topics', "
            f"borrower:'{bo_ct} entities &middot; {bo_tp} topics', "
            f"bank:'{ba_ct} entities &middot; {ba_tp} topics' }}"
        )
        audit_data_js = (
            "{\n  lender: "
            + json.dumps(lender_data["audit_docs"], default=str)
            + ",\n  borrower: "
            + json.dumps(borrower_data["audit_docs"], default=str)
            + ",\n  bank: "
            + json.dumps(bank_data["audit_docs"], default=str)
            + "\n}"
        )
        charts_flags_js = "{ lender:false, borrower:false, bank:false }"
        audits_flags_js = "{ lender:false, borrower:false, bank:false }"
        bank_chart_block = (
            f"  if (layer === 'bank') {{\n"
            f"    const bs = {json.dumps(bank_data['scores'])};\n"
            "    chartInstances['bankChart'] = new Chart(document.getElementById('bankChart'), { type:'bar', data:{ labels:"
            + json.dumps(bank_data["labels"])
            + ", datasets:[{ label:'Net Position', data:bs, backgroundColor:bs.map(v=>v>0?'#00d4aa':'#FF6B6B'), borderRadius:4 }] }, options:{ indexAxis:'y', responsive:true, plugins:{legend:{display:false}}, scales:{ x:{grid:{color:'#21262d'}}, y:{grid:{display:false}} } } } });\n"
            "    buildHeatmap('bankHeatmap', 'bank', "
            + json.dumps(bank_data["heatmap_entities"])
            + ", "
            + json.dumps(bank_data["heatmap_topics"])
            + ", "
            + json.dumps(bank_data["heatmap_data"])
            + ", "
            + json.dumps(bank_data["heatmap_topic_polarities"])
            + ");\n"
            "  }\n"
        )
    else:
        bank_page_html = ""
        bank_nav_html = ""
        page_titles_js = (
            "{ overview:'Private Credit Stress Analyzer', lender:'Lender Terms Power Analysis', "
            "borrower:'Borrower Distress Analysis' }"
        )
        page_badges_js = (
            f"{{ overview:'Overview &middot; thematic demo', lender:'{le_ct} entities &middot; {le_tp} topics', "
            f"borrower:'{bo_ct} entities &middot; {bo_tp} topics' }}"
        )
        audit_data_js = (
            "{\n  lender: "
            + json.dumps(lender_data["audit_docs"], default=str)
            + ",\n  borrower: "
            + json.dumps(borrower_data["audit_docs"], default=str)
            + "\n}"
        )
        charts_flags_js = "{ lender:false, borrower:false }"
        audits_flags_js = "{ lender:false, borrower:false }"
        bank_chart_block = ""

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Private Credit Stress Analyzer — Bigdata</title>
<link rel="icon" type="image/svg+xml" href="{favicon_svg}">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
html, body {{ height:100%; }}
body {{ background:#0d1117; color:#c9d1d9; font-family:'Inter',system-ui,-apple-system,sans-serif; display:flex; height:100vh; max-height:100vh; overflow:hidden; }}

.sidebar {{ width:220px; min-width:220px; background:#161b22; border-right:1px solid #30363d; display:flex; flex-direction:column; }}
.sidebar-logo {{ padding:1.1rem 1.2rem; display:flex; align-items:center; gap:0.6rem; border-bottom:1px solid #30363d; }}
.sidebar-logo svg {{ width:26px; height:26px; flex-shrink:0; }}
.sidebar-logo span {{ font-size:1.05rem; font-weight:700; color:#fff; }}
.sidebar-logo .dot {{ color:#8b949e; font-weight:400; font-size:0.82rem; }}
.sidebar-section {{ padding:1rem 0 0; }}
.sidebar-section-label {{ padding:0 1.2rem; font-size:0.62rem; text-transform:uppercase; letter-spacing:1.5px; color:#484f58; font-weight:600; margin-bottom:0.5rem; }}
.nav-item {{ display:flex; align-items:center; gap:0.7rem; padding:0.6rem 1.2rem; cursor:pointer; color:#8b949e; font-size:0.85rem; font-weight:500; transition:all 0.15s; border-left:3px solid transparent; }}
.nav-item:hover {{ background:#1c2128; color:#c9d1d9; }}
.nav-item.active {{ background:rgba(76,167,249,0.08); color:#4CA7F9; border-left-color:#4CA7F9; }}
.nav-item svg {{ width:18px; height:18px; flex-shrink:0; opacity:0.7; }}
.nav-item.active svg {{ opacity:1; }}
.sidebar-footer {{ margin-top:auto; padding:1rem 1.2rem; border-top:1px solid #30363d; font-size:0.68rem; color:#484f58; line-height:1.5; }}

/* One vertical scroll: .content only (.main does not scroll — avoids nested scrollbars). */
.main {{ flex:1; min-width:0; min-height:0; overflow:hidden; display:flex; flex-direction:column; }}
.topbar {{ flex-shrink:0; padding:1rem 2rem; border-bottom:1px solid #30363d; display:flex; align-items:center; justify-content:space-between; background:#161b22; min-height:56px; }}
.topbar h1 {{ font-size:1.15rem; font-weight:600; color:#fff; }}
.topbar .badge {{ background:rgba(76,167,249,0.15); color:#4CA7F9; font-size:0.7rem; padding:0.2rem 0.6rem; border-radius:99px; font-weight:600; }}
.content {{ flex:1; min-height:0; padding:1.5rem 2rem 2rem; overflow-y:auto; overflow-x:hidden; }}

.tabs {{ display:flex; gap:0; border-bottom:1px solid #30363d; margin-bottom:1.5rem; }}
.tab {{ padding:0.65rem 1.2rem; font-size:0.8rem; font-weight:500; color:#8b949e; cursor:pointer; border-bottom:2px solid transparent; transition:all 0.15s; }}
.tab:hover {{ color:#c9d1d9; }}
.tab.active {{ color:#00d4aa; border-bottom-color:#00d4aa; }}
.tab-panel {{ display:none; }}
.tab-panel.active {{ display:block; }}

.card {{ background:#161b22; border:1px solid #30363d; border-radius:8px; padding:1.5rem; margin-bottom:1.5rem; }}
.card h3 {{ color:#00d4aa; font-size:1rem; margin-bottom:1rem; font-weight:600; }}
canvas {{ max-height:420px; }}
.borrower-score-chart-wrap, .lender-score-chart-wrap {{ position:relative; width:100%; min-height:320px; }}
.borrower-score-chart-wrap canvas, .lender-score-chart-wrap canvas {{ max-height:none !important; height:auto !important; }}

.heatmap-wrap {{ overflow-x:auto; }}
.hm {{ border-collapse:collapse; width:100%; font-size:0.75rem; }}
.hm th {{ background:#1a1a2e; color:#00d4aa; padding:8px 10px; text-align:center; font-weight:600; white-space:nowrap; position:sticky; top:0; z-index:1; vertical-align:bottom; }}
.hm th.hm-th-neg {{ color:#FF6B6B; }}
.hm td {{ padding:7px 10px; text-align:center; border:1px solid #21262d; }}
.hm td.hm-cell {{ cursor:pointer; transition:filter 0.12s ease, box-shadow 0.12s ease; }}
.hm td.hm-cell:hover {{ filter:brightness(1.12); box-shadow:inset 0 0 0 1px rgba(255,255,255,0.35); }}
.hm td.hm-cell:focus {{ outline:2px solid #00d4aa; outline-offset:-2px; }}
.hm tr:nth-child(even) {{ background:#0d1117; }}
.hm tr:nth-child(odd) {{ background:#161b22; }}

.themes-list {{ display:flex; flex-direction:column; gap:0.5rem; }}
.theme-card {{ padding:0.75rem 1rem; border-radius:6px; }}
.theme-card.positive {{ background:rgba(0,212,170,0.06); border-left:3px solid #00d4aa; }}
.theme-card.negative {{ background:rgba(255,107,107,0.06); border-left:3px solid #FF6B6B; }}
.theme-header {{ display:flex; align-items:center; gap:0.6rem; }}
.theme-name {{ font-size:0.85rem; font-weight:600; color:#c9d1d9; flex:1; display:flex; align-items:center; gap:0.35rem; flex-wrap:wrap; min-width:0; }}
.theme-count {{ font-weight:700; font-size:1rem; min-width:2.5rem; text-align:right; }}
.theme-card.positive .theme-count {{ color:#00d4aa; }}
.theme-card.negative .theme-count {{ color:#FF6B6B; }}
.theme-query {{ font-size:0.75rem; color:#8b949e; font-style:italic; margin-top:0.35rem; padding-left:0.1rem; }}
.theme-query-label {{ font-style:normal; font-weight:600; color:#c9d1d9; }}
.theme-query em {{ color:#c9d1d9; font-style:normal; font-weight:600; }}
.pol-badge {{ font-size:0.65rem; padding:0.1rem 0.45rem; border-radius:99px; font-weight:600; white-space:nowrap; }}
.pol-badge.pos {{ background:rgba(0,212,170,0.15); color:#00d4aa; }}
.pol-badge.neg {{ background:rgba(255,107,107,0.15); color:#FF6B6B; }}

.audit-filters {{ display:flex; gap:0.75rem; margin-bottom:1rem; flex-wrap:wrap; }}
.audit-filters select {{ background:#0d1117; color:#c9d1d9; border:1px solid #30363d; border-radius:6px; padding:0.4rem 0.7rem; font-size:0.8rem; cursor:pointer; min-width:160px; }}
.audit-filters select:focus {{ border-color:#00d4aa; outline:none; }}
.audit-count {{ font-size:0.75rem; color:#8b949e; margin-bottom:0.75rem; }}
/* Table grows with page; .content is the only vertical scroller (sticky thead uses .content as scroll root). */
.audit-table-wrap {{ overflow-x:auto; }}
.at {{ border-collapse:collapse; width:100%; font-size:0.78rem; }}
.at th {{ background:#1a1a2e; color:#00d4aa; padding:8px 10px; text-align:left; font-weight:600; white-space:nowrap; position:sticky; top:0; z-index:1; }}
.at td {{ padding:8px 10px; border-bottom:1px solid #21262d; vertical-align:top; }}
.at tr:hover {{ background:#1c2128; }}
.at .td-hl {{ font-weight:600; color:#c9d1d9; max-width:280px; }}
.at .td-snip {{ color:#8b949e; font-size:0.73rem; max-width:350px; line-height:1.4; }}
.at .td-link a {{ color:#4CA7F9; text-decoration:none; font-size:0.72rem; }}
.at .td-link a:hover {{ text-decoration:underline; }}
.at .td-date {{ white-space:nowrap; color:#8b949e; }}

.method-block {{ font-size:0.85rem; line-height:1.7; color:#8b949e; }}
.method-block h4 {{ color:#c9d1d9; font-size:0.9rem; margin:1.2rem 0 0.3rem; }}
.method-block code {{ background:#1a1a2e; padding:0.15rem 0.4rem; border-radius:4px; color:#00d4aa; font-size:0.8rem; }}

.method-hero {{ text-align:center; margin-bottom:2rem; }}
.method-hero h2 {{ color:#fff; font-size:1.35rem; font-weight:700; margin-bottom:0.5rem; }}
.method-hero p {{ color:#8b949e; max-width:600px; margin:0 auto; font-size:0.88rem; line-height:1.6; }}
.method-hero strong {{ color:#c9d1d9; }}

.method-cards {{ display:grid; grid-template-columns:1fr 1fr; gap:1rem; margin-bottom:2rem; }}
.method-card {{ background:#161b22; border:1px solid #30363d; border-radius:10px; padding:1.3rem 1.4rem; }}
.method-card h3 {{ color:#4CA7F9; font-size:0.92rem; font-weight:700; margin-bottom:0.55rem; display:flex; align-items:center; gap:0.45rem; }}
.method-card p {{ color:#8b949e; font-size:0.82rem; line-height:1.6; }}

.method-steps {{ position:relative; padding-left:2.6rem; }}
.method-step {{ position:relative; margin-bottom:1.6rem; }}
.method-step:last-child {{ margin-bottom:0; }}
.method-step::before {{ content:''; position:absolute; left:-1.8rem; top:1.6rem; bottom:-1.6rem; width:2px; background:#30363d; }}
.method-step:last-child::before {{ display:none; }}
.method-step-num {{ position:absolute; left:-2.6rem; top:0; width:1.7rem; height:1.7rem; border-radius:50%; display:flex; align-items:center; justify-content:center; font-size:0.72rem; font-weight:700; color:#fff; z-index:1; }}
.method-step-num.blue {{ background:#4CA7F9; }}
.method-step-num.green {{ background:#00d4aa; }}
.method-step-num.amber {{ background:#FFD93D; color:#0d1117; }}
.method-step-num.red {{ background:#FF6B6B; }}
.method-step h4 {{ color:#fff; font-size:0.95rem; margin:0 0 0.35rem; font-weight:600; }}
.method-step p {{ color:#8b949e; font-size:0.82rem; line-height:1.55; margin:0; }}
.method-step .step-detail {{ background:#0d1117; border:1px solid #21262d; border-radius:6px; padding:0.6rem 0.9rem; margin-top:0.5rem; font-size:0.78rem; color:#4CA7F9; font-style:italic; }}
.method-step code {{ background:#1a1a2e; padding:0.12rem 0.35rem; border-radius:4px; color:#00d4aa; font-size:0.78rem; }}

.stats {{ display:flex; gap:1rem; margin-bottom:1.5rem; }}
.stat-card {{ flex:1; background:#161b22; border:1px solid #30363d; border-radius:8px; padding:1rem 1.2rem; text-align:center; }}
.stat-val {{ font-size:1.6rem; font-weight:700; color:#00d4aa; }}
.stat-label {{ font-size:0.72rem; color:#8b949e; text-transform:uppercase; letter-spacing:0.5px; margin-top:0.2rem; }}
.stat-desc {{ font-size:0.7rem; color:#484f58; line-height:1.3; margin-top:0.35rem; max-width:140px; margin-left:auto; margin-right:auto; }}

.card-desc {{ font-size:0.78rem; color:#8b949e; margin-bottom:0.75rem; }}

.overview-wrap {{ max-width:780px; }}
.overview-hero {{ margin-bottom:1.5rem; }}
.overview-hero h2 {{ color:#fff; font-size:1.25rem; margin-bottom:0.75rem; font-weight:700; }}
.overview-lead {{ color:#c9d1d9; font-size:0.88rem; line-height:1.6; margin-bottom:0.5rem; }}
.overview-lead a {{ color:#4CA7F9; text-decoration:none; }}
.overview-lead a:hover {{ text-decoration:underline; }}
.overview-list {{ margin:0.5rem 0 0 1.1rem; color:#8b949e; font-size:0.85rem; line-height:1.65; }}
.overview-list li {{ margin-bottom:0.35rem; }}

.overview-section {{ margin-bottom:1.5rem; }}
.overview-section h3 {{ color:#fff; font-size:1rem; font-weight:600; margin-bottom:0.6rem; }}
.overview-config-list {{ list-style:none; margin:0; padding:0; display:flex; flex-direction:column; gap:0.5rem; }}
.overview-config-list li {{ display:flex; gap:0.5rem; font-size:0.84rem; color:#8b949e; line-height:1.55; padding:0.6rem 0.85rem; background:#161b22; border:1px solid #30363d; border-radius:6px; }}
.config-label {{ font-weight:600; color:#4CA7F9; white-space:nowrap; min-width:5.5rem; flex-shrink:0; }}

.overview-disclaimer {{ padding:1rem 1.15rem; background:rgba(255,217,61,0.06); border:1px solid rgba(255,217,61,0.18); border-radius:8px; }}
.overview-disclaimer h3 {{ color:#FFD93D; }}
.overview-disclaimer p {{ color:#c9d1d9; font-size:0.84rem; line-height:1.6; margin:0; }}
.overview-disclaimer a {{ color:#4CA7F9; text-decoration:none; }}
.overview-disclaimer a:hover {{ text-decoration:underline; }}

.overview-scores-heading {{ color:#fff; font-size:1.1rem; font-weight:700; margin:0 0 0.35rem; }}
.overview-scores-lead {{ color:#8b949e; font-size:0.84rem; line-height:1.55; margin-bottom:1rem; }}
.overview-scores-grid {{ display:grid; grid-template-columns:1fr 1fr; gap:1rem; margin-bottom:1rem; }}
@media (max-width:700px) {{ .overview-scores-grid {{ grid-template-columns:1fr; }} }}
.overview-card {{ background:#161b22; border:1px solid #30363d; border-radius:8px; padding:1.25rem 1.4rem; }}
.overview-card h3 {{ color:#00d4aa; font-size:0.92rem; margin-bottom:0.65rem; font-weight:600; }}
.overview-card p {{ color:#8b949e; font-size:0.82rem; line-height:1.6; margin-bottom:0.5rem; }}
.overview-card p:last-child {{ margin-bottom:0; }}
.overview-card code {{ display:block; background:#0d1117; padding:0.45rem 0.6rem; border-radius:6px; color:#00d4aa; font-size:0.78rem; margin:0.35rem 0; }}
.overview-footer {{ font-size:0.78rem; color:#484f58; line-height:1.5; margin-top:0.75rem; }}

.tab-intro {{ font-size:0.8rem; color:#8b949e; line-height:1.55; margin:0 0 1rem 0; max-width:900px; }}
.layer-blurb {{ font-size:0.82rem; color:#c9d1d9; line-height:1.55; margin:0 0 1rem 0; max-width:880px; }}

.theme-desc {{ font-weight:400; font-size:0.75rem; color:#8b949e; }}


.layer-page {{ display:none; }}
.layer-page.active {{ display:block; }}

/* ── Auth / Settings overlay ─────────────────────────────── */
.overlay {{ position:fixed; inset:0; z-index:3000; background:rgba(0,0,0,0.7); display:none; align-items:center; justify-content:center; }}
.overlay.visible {{ display:flex; }}
.overlay-box {{ background:#161b22; border:1px solid #30363d; border-radius:12px; padding:28px; width:90%; max-width:440px; }}
.overlay-box h2 {{ color:#fff; font-size:1.1rem; margin-bottom:0.5rem; display:flex; align-items:center; gap:0.5rem; }}
.overlay-box p.overlay-desc {{ color:#8b949e; font-size:0.82rem; line-height:1.55; margin-bottom:1rem; }}
.api-key-input-wrap {{ display:flex; gap:0.5rem; margin-bottom:0.75rem; }}
.api-key-input-wrap input {{ flex:1; background:#0d1117; color:#c9d1d9; border:1px solid #30363d; border-radius:6px; padding:0.5rem 0.7rem; font-size:0.85rem; }}
.api-key-input-wrap input:focus {{ border-color:#4CA7F9; outline:none; }}
.eye-btn {{ background:none; border:none; color:#8b949e; cursor:pointer; font-size:1rem; padding:0 0.3rem; }}
#apiKeyStatus {{ font-size:0.75rem; color:#8b949e; margin-bottom:0.75rem; }}
.overlay-actions {{ display:flex; gap:0.5rem; justify-content:flex-end; }}
.overlay-actions button {{ padding:0.45rem 1rem; border-radius:6px; border:1px solid #30363d; font-size:0.8rem; font-weight:600; cursor:pointer; }}
.btn-save-key {{ background:#4CA7F9; color:#fff; border-color:#4CA7F9; }}
.btn-save-key:hover {{ background:#3d96e0; }}
.btn-clear-key {{ background:transparent; color:#FF6B6B; border-color:#FF6B6B; }}
.btn-close-settings {{ background:transparent; color:#8b949e; }}
.nav-item-settings {{ margin-top:0.5rem; border-top:1px solid #21262d; padding-top:0.5rem; }}

/* ── Customize Entity overlay ────────────────────────────── */
.entity-overlay {{ max-width:700px; }}
.entity-overlay textarea {{ width:100%; background:#0d1117; color:#c9d1d9; border:1px solid #30363d; border-radius:6px; padding:0.6rem; font-size:0.82rem; min-height:80px; resize:vertical; font-family:inherit; }}
.entity-overlay textarea:focus {{ border-color:#4CA7F9; outline:none; }}
.entity-table {{ width:100%; border-collapse:collapse; font-size:0.8rem; margin:0.75rem 0; }}
.entity-table th {{ background:#1a1a2e; color:#4CA7F9; padding:8px 10px; text-align:left; font-weight:600; }}
.entity-table td {{ padding:8px 10px; border-bottom:1px solid #21262d; }}
.entity-table tr:hover {{ background:#1c2128; }}
.btn-remove-entity {{ background:none; border:none; color:#FF6B6B; cursor:pointer; font-size:0.8rem; font-weight:600; }}
.btn-remove-entity:hover {{ text-decoration:underline; }}
.entity-step {{ display:none; }}
.entity-step.active {{ display:block; }}

/* ── Pipeline progress overlay ───────────────────────────── */
.pipeline-overlay {{ max-width:400px; text-align:center; }}
.pipeline-spinner {{ width:40px; height:40px; border:3px solid #30363d; border-top-color:#4CA7F9; border-radius:50%; animation:spin 0.8s linear infinite; margin:0 auto 1rem; }}
@keyframes spin {{ to {{ transform:rotate(360deg); }} }}
#pipelineProgress {{ color:#c9d1d9; font-size:0.85rem; margin-bottom:0.5rem; }}
#pipelineElapsed {{ color:#484f58; font-size:0.75rem; }}

/* ── Topbar customize button ─────────────────────────────── */
.topbar-right {{ display:flex; align-items:center; gap:0.6rem; }}
.btn-customize {{ background:rgba(0,212,170,0.12); color:#00d4aa; border:1px solid rgba(0,212,170,0.3); border-radius:6px; padding:0.3rem 0.8rem; font-size:0.72rem; font-weight:600; cursor:pointer; display:none; white-space:nowrap; }}
.btn-customize:hover {{ background:rgba(0,212,170,0.2); }}
.btn-reset-custom {{ background:rgba(255,107,107,0.12); color:#FF6B6B; border:1px solid rgba(255,107,107,0.3); border-radius:6px; padding:0.3rem 0.8rem; font-size:0.72rem; font-weight:600; cursor:pointer; display:none; white-space:nowrap; }}
.btn-reset-custom:hover {{ background:rgba(255,107,107,0.2); }}
</style>
</head>
<body>

<!-- ── API Key Gate Overlay ────────────────────────────── -->
<div class="overlay" id="settingsOverlay">
  <div class="overlay-box">
    <h2>&#128273; Bigdata API Key</h2>
    <p class="overlay-desc" id="settingsApiKeyDescription">Enter your Bigdata API key to continue. The dashboard will be available after you save.</p>
    <div class="api-key-input-wrap">
      <input type="password" id="settingsApiKey" placeholder="Paste your Bigdata API key">
      <button class="eye-btn" onclick="toggleApiKeyVisibility()" title="Show / hide key">&#128065;</button>
    </div>
    <div id="apiKeyStatus"></div>
    <div class="overlay-actions">
      <button class="btn-close-settings" onclick="closeSettings()">Cancel</button>
      <button class="btn-clear-key" onclick="clearApiKey()">Clear</button>
      <button class="btn-save-key" onclick="saveApiKey()">Save</button>
    </div>
  </div>
</div>

<!-- ── Custom Entity Overlay ──────────────────────────── -->
<div class="overlay" id="entityOverlay">
  <div class="overlay-box entity-overlay">
    <div class="entity-step active" id="entityStepInput">
      <h2>Customize Entities</h2>
      <p class="overlay-desc">Enter one company per line (best for names with commas), or comma-separated names, for the <strong id="entityLayerLabel">borrower</strong> layer.</p>
      <textarea id="entityNamesInput" placeholder="Chegg&#10;2U, Inc.&#10;Pluralsight"></textarea>
      <div class="overlay-actions" style="margin-top:0.75rem;">
        <button class="btn-close-settings" onclick="closeEntityOverlay()">Cancel</button>
        <button class="btn-save-key" onclick="lookupEntities()">Look Up</button>
      </div>
    </div>
    <div class="entity-step" id="entityStepConfirm">
      <h2>Confirm Entities</h2>
      <p class="overlay-desc">Review the resolved companies below. Remove any you don't want, then confirm.</p>
      <div style="max-height:350px;overflow-y:auto;" id="entityTableWrap"></div>
      <div class="overlay-actions" style="margin-top:0.75rem;">
        <button class="btn-close-settings" onclick="backToEntityInput()">Back</button>
        <button class="btn-close-settings" onclick="closeEntityOverlay()">Cancel</button>
        <button class="btn-save-key" onclick="confirmEntities()">Confirm &amp; Run</button>
      </div>
    </div>
  </div>
</div>

<!-- ── Pipeline Progress Overlay ──────────────────────── -->
<div class="overlay" id="pipelineOverlay">
  <div class="overlay-box pipeline-overlay">
    <div class="pipeline-spinner"></div>
    <h2>Running Pipeline</h2>
    <div id="pipelineProgress">Starting...</div>
    <div id="pipelineElapsed"></div>
  </div>
</div>

<div id="appContent" style="display:none;flex:1;min-width:0;min-height:0;height:100vh;">
<nav class="sidebar">
  <div class="sidebar-logo">
    <svg viewBox="0 0 159.7 159.7"><path fill="#4CA7F9" d="M38.11,0h83.48c21.03,0,38.11,17.08,38.11,38.11v83.48c0,21.03-17.08,38.11-38.11,38.11H38.11c-21.03,0-38.11-17.08-38.11-38.11V38.11C0,17.08,17.08,0,38.11,0Z"/><path fill="#FFFDF5" d="M105.69,137.06c-8.4,0-16.35-3.27-22.4-9.21-6.07-5.96-9.41-13.84-9.41-22.18v-51.63c0-11.38-8.72-20.3-19.87-20.3-5.4,0-10.44,2.12-14.21,5.96-3.74,3.82-5.81,8.91-5.81,14.34s2.06,10.52,5.81,14.34c3.76,3.84,8.81,5.96,14.21,5.96h13.36v11.09h-15.49c-8.26,0-15.83-3.26-21.32-9.19-5.4-5.83-8.37-13.71-8.37-22.2s3.34-16.22,9.41-22.18c6.05-5.94,14-9.21,22.4-9.21s16.36,3.27,22.4,9.21c6.07,5.96,9.41,13.84,9.41,22.18v51.63c0,11.38,8.72,20.3,19.86,20.3,5.4,0,10.44-2.12,14.21-5.96,3.74-3.82,5.81-8.91,5.81-14.34s-2.06-10.52-5.81-14.34c-3.76-3.84-8.81-5.96-14.21-5.96h-13.26v-11.09h15.4c8.26,0,15.83,3.26,21.32,9.19,5.4,5.82,8.37,13.71,8.37,22.2s-3.34,16.22-9.41,22.18c-6.05,5.94-14,9.21-22.4,9.21Z"/></svg>
    <span>bigdata<span class="dot">.com</span></span>
  </div>
  <div class="sidebar-section">
    <div class="sidebar-section-label">Analysis Layers</div>
    <div class="nav-item active" data-layer="overview" onclick="switchLayer('overview')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M3 9l9-7 9 7v11a2 2 0 0 1-2 2H5a2 2 0 0 1-2-2z"/><polyline points="9 22 9 12 15 12 15 22"/></svg>
      Overview
    </div>
    <div class="nav-item" data-layer="borrower" onclick="switchLayer('borrower')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><path d="M12 2L2 7l10 5 10-5-10-5z"/><path d="M2 17l10 5 10-5"/><path d="M2 12l10 5 10-5"/></svg>
      Borrowers
    </div>
    <div class="nav-item" data-layer="lender" onclick="switchLayer('lender')">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="3" width="7" height="7" rx="1"/><rect x="14" y="3" width="7" height="7" rx="1"/><rect x="3" y="14" width="7" height="7" rx="1"/><rect x="14" y="14" width="7" height="7" rx="1"/></svg>
      Lenders
    </div>{bank_nav_html}
    <div class="nav-item nav-item-settings" onclick="openSettings()">
      <svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><rect x="3" y="11" width="18" height="11" rx="2" ry="2"/><path d="M7 11V7a5 5 0 0 1 10 0v4"/></svg>
      API Key
    </div>
  </div>
  <div class="sidebar-footer">Private Credit Stress Analyzer<br>Powered by Bigdata API</div>
</nav>

<div class="main">
  <div class="topbar">
    <h1 id="pageTitle">Private Credit Stress Analyzer</h1>
    <div class="topbar-right">
      <button class="btn-customize" id="btnCustomize" onclick="openCustomEntityModal()">Customize Entities</button>
      <button class="btn-reset-custom" id="btnResetCustom" onclick="resetToDefault()">Reset to Default</button>
      <span class="badge" id="pageBadge">Overview &middot; thematic demo</span>
    </div>
  </div>
  <div class="content">
    {overview_page}
    {lender_page}
    {borrower_page}
    {bank_page_html}
  </div>
</div>
</div><!-- /appContent -->

<script>
Chart.defaults.color = '#c9d1d9';
Chart.defaults.borderColor = '#30363d';

const pageTitles = {page_titles_js};
const pageBadges = {page_badges_js};

const auditData = {audit_data_js};

let chartsInitialized = {charts_flags_js};
let auditsInitialized = {audits_flags_js};

function switchLayer(layer) {{
  document.querySelectorAll('.nav-item').forEach(n => n.classList.remove('active'));
  document.querySelector('.nav-item[data-layer="'+layer+'"]').classList.add('active');
  document.querySelectorAll('.layer-page').forEach(p => p.classList.remove('active'));
  document.getElementById('page-'+layer).classList.add('active');
  document.getElementById('pageTitle').textContent = pageTitles[layer];
  document.getElementById('pageBadge').innerHTML = pageBadges[layer];
  if (layer !== 'overview') {{
    if (!chartsInitialized[layer]) initCharts(layer);
    if (!auditsInitialized[layer]) initAudit(layer);
  }}
}}

function switchTabTo(layer, tab) {{
  const page = document.getElementById('page-'+layer);
  if (!page) return;
  page.querySelectorAll('.tab').forEach(t => t.classList.remove('active'));
  page.querySelectorAll('.tab-panel').forEach(p => p.classList.remove('active'));
  const tabBtn = page.querySelector('.tab[data-tab="'+tab+'"]');
  if (tabBtn) tabBtn.classList.add('active');
  const panel = document.getElementById(layer+'-'+tab);
  if (panel) panel.classList.add('active');
  if (tab === 'audit' && !auditsInitialized[layer]) initAudit(layer);
}}

function switchTab(layer, tab) {{
  switchTabTo(layer, tab);
}}

function escapeHeatmapAttr(s) {{
  return String(s).replace(/&/g,'&amp;').replace(/"/g,'&quot;').replace(/</g,'&lt;');
}}

function escapeHtml(s) {{
  return String(s == null ? '' : s)
    .replace(/&/g,'&amp;')
    .replace(/</g,'&lt;')
    .replace(/>/g,'&gt;')
    .replace(/"/g,'&quot;');
}}

/** Only http(s) URLs for links; blocks javascript: and other schemes. */
function safeHttpUrl(u) {{
  if (u == null || typeof u !== 'string') return '';
  const t = u.trim();
  if (!t) return '';
  const low = t.slice(0, 8).toLowerCase();
  if (low.startsWith('https://') || low.startsWith('http://')) return t;
  return '';
}}

function openAuditFromHeatmap(layer, entity, topic) {{
  switchLayer(layer);
  const selE = document.getElementById(layer+'FEntity');
  const selT = document.getElementById(layer+'FTopic');
  const selP = document.getElementById(layer+'FPol');
  if (selE) {{
    let ok = false;
    for (let k = 0; k < selE.options.length; k++) {{
      if (selE.options[k].value === entity) {{ selE.selectedIndex = k; ok = true; break; }}
    }}
    if (!ok && entity) {{
      const o = document.createElement('option');
      o.value = entity; o.textContent = entity;
      selE.appendChild(o);
      selE.value = entity;
    }}
  }}
  if (selT) {{
    let ok = false;
    for (let k = 0; k < selT.options.length; k++) {{
      if (selT.options[k].value === topic) {{ selT.selectedIndex = k; ok = true; break; }}
    }}
    if (!ok && topic) {{
      const o = document.createElement('option');
      o.value = topic; o.textContent = topic;
      selT.appendChild(o);
      selT.value = topic;
    }}
  }}
  if (selP) selP.value = '';
  switchTabTo(layer, 'audit');
  renderAudit(layer);
}}

function buildHeatmap(id, layer, entities, topics, data, polarities) {{
  let mx = 1;
  data.forEach(r => r.forEach(v => {{ if(v>mx) mx=v; }}));
  let h = '<table class="hm"><thead><tr><th>Entity</th>';
  topics.forEach((t, j) => {{
    const neg = polarities && polarities[j] === 'negative';
    h += '<th'+(neg ? ' class="hm-th-neg"' : '')+'>'+escapeHtml(t)+'</th>';
  }});
  h += '</tr></thead><tbody>';
  data.forEach((row,i) => {{
    h += '<tr><td style="text-align:left;font-weight:600;white-space:nowrap;">'+escapeHtml(entities[i])+'</td>';
    row.forEach((v, j) => {{
      const p = Math.min(v/mx,1);
      const neg = polarities && polarities[j] === 'negative';
      let bg;
      if (neg) {{
        const r = Math.round(45 + p * (255 - 45));
        const g = Math.round(27 + p * (107 - 27));
        const b = Math.round(27 + p * (107 - 27));
        bg = 'rgba('+r+','+g+','+b+','+(0.12+p*0.88)+')';
      }} else {{
        bg = 'rgba('+Math.round(13+p*(0-13))+','+Math.round(17+p*(212-17))+','+Math.round(23+p*(170-23))+','+(0.15+p*0.85)+')';
      }}
      const entA = escapeHeatmapAttr(entities[i]);
      const topA = escapeHeatmapAttr(topics[j]);
      h += '<td class="hm-cell" tabindex="0" role="button" title="Open audit for this entity and topic" data-layer="'+layer+'" data-entity="'+entA+'" data-topic="'+topA+'" style="background:'+bg+';color:#fff;">'+v+'</td>';
    }});
    h += '</tr>';
  }});
  h += '</tbody></table>';
  const el = document.getElementById(id);
  el.innerHTML = h;
  el.onclick = function(ev) {{
    const td = ev.target.closest('td.hm-cell');
    if (!td) return;
    const L = td.getAttribute('data-layer');
    const ent = td.getAttribute('data-entity');
    const top = td.getAttribute('data-topic');
    if (L && ent != null && top != null) openAuditFromHeatmap(L, ent, top);
  }};
  el.onkeydown = function(ev) {{
    if (ev.key !== 'Enter' && ev.key !== ' ') return;
    const td = ev.target.closest('td.hm-cell');
    if (!td) return;
    ev.preventDefault();
    const L = td.getAttribute('data-layer');
    const ent = td.getAttribute('data-entity');
    const top = td.getAttribute('data-topic');
    if (L && ent != null && top != null) openAuditFromHeatmap(L, ent, top);
  }};
}}

function initAudit(layer) {{
  auditsInitialized[layer] = true;
  const docs = auditData[layer];
  if (!docs || !docs.length) return;

  const entities = [...new Set(docs.map(d => d.entity))].sort();
  const topics = [...new Set(docs.map(d => d.topic))].sort();
  const polarities = ['positive','negative'];

  const fc = document.getElementById(layer+'AuditFilters');
  fc.innerHTML = '<select id="'+layer+'FEntity"><option value="">All Entities</option>'+entities.map(e=>'<option value="'+escapeHtml(e)+'">'+escapeHtml(e)+'</option>').join('')+'</select>'
    + '<select id="'+layer+'FTopic"><option value="">All Topics</option>'+topics.map(t=>'<option value="'+escapeHtml(t)+'">'+escapeHtml(t)+'</option>').join('')+'</select>'
    + '<select id="'+layer+'FPol"><option value="">All Polarities</option>'+polarities.map(p=>{{ const lab = p[0].toUpperCase()+p.slice(1); return '<option value="'+escapeHtml(p)+'">'+escapeHtml(lab)+'</option>'; }}).join('')+'</select>';

  [layer+'FEntity',layer+'FTopic',layer+'FPol'].forEach(id => {{
    document.getElementById(id).addEventListener('change', () => renderAudit(layer));
  }});
  renderAudit(layer);
}}

function renderAudit(layer) {{
  const docs = auditData[layer] || [];
  const entEl = document.getElementById(layer+'FEntity');
  const topEl = document.getElementById(layer+'FTopic');
  const polEl = document.getElementById(layer+'FPol');
  if (!entEl || !topEl || !polEl) return;
  const ent = entEl.value;
  const top = topEl.value;
  const pol = polEl.value;

  let filtered = docs;
  if (ent) filtered = filtered.filter(d => d.entity === ent);
  if (top) filtered = filtered.filter(d => d.topic === top);
  if (pol) filtered = filtered.filter(d => d.polarity === pol);

  document.getElementById(layer+'AuditCount').textContent = 'Showing '+filtered.length+' of '+docs.length+' documents';

  let h = '<table class="at"><thead><tr><th>Entity</th><th>Topic</th><th>Polarity</th><th>Headline</th><th>Content</th><th>Date</th><th>Source</th></tr></thead><tbody>';
  filtered.slice(0,200).forEach(d => {{
    const pc = d.polarity === 'positive' ? 'pos' : 'neg';
    const pi = d.polarity === 'positive' ? '+' : '&minus;';
    const href = safeHttpUrl(d.url);
    const link = href ? '<a href="'+escapeHtml(href)+'" target="_blank" rel="noopener noreferrer">View &rarr;</a>' : '';
    const pol = String(d.polarity || '');
    h += '<tr><td style="font-weight:600;white-space:nowrap;">'+escapeHtml(d.entity)+'</td>'
      + '<td style="white-space:nowrap;">'+escapeHtml(d.topic)+'</td>'
      + '<td><span class="pol-badge '+pc+'">'+pi+' '+escapeHtml(pol)+'</span></td>'
      + '<td class="td-hl">'+escapeHtml(d.headline)+'</td>'
      + '<td class="td-snip">'+escapeHtml(d.snippet)+'</td>'
      + '<td class="td-date">'+escapeHtml(d.timestamp)+'</td>'
      + '<td class="td-link">'+link+'</td></tr>';
  }});
  h += '</tbody></table>';
  document.getElementById(layer+'AuditTable').innerHTML = h;
}}

function initCharts(layer) {{
  chartsInitialized[layer] = true;
  if (layer === 'lender') {{
    chartInstances['lenderRadarChart'] = new Chart(document.getElementById('lenderRadarChart'), {{ type:'radar', data:{{ labels:{lender_radar_labels}, datasets:{lender_radar_datasets} }}, options:{{ responsive:true, plugins:{{legend:{{position:'bottom',labels:{{boxWidth:12}}}}}}, scales:{{ r:{{beginAtZero:true, grid:{{color:'#21262d'}}, angleLines:{{color:'#21262d'}}, pointLabels:{{font:{{size:10}},color:'#c9d1d9'}}}} }} }} }});
    const s = {json.dumps(lender_data["scores"])};
    const lLabels = {json.dumps(lender_data["labels"])};
    const lScoreCanvas = document.getElementById('lenderChart');
    const lScoreWrap = lScoreCanvas && lScoreCanvas.closest('.lender-score-chart-wrap');
    if (lScoreWrap) {{ lScoreWrap.style.height = Math.max(320, lLabels.length * 28) + 'px'; }}
    chartInstances['lenderChart'] = new Chart(lScoreCanvas, {{ type:'bar', data:{{ labels:lLabels, datasets:[{{ label:'Terms Power Score', data:s, backgroundColor:s.map(v=>v>50?'#00d4aa':v>30?'#FFD93D':'#FF6B6B'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}, ticks:{{autoSkip:false, maxRotation:0}}}} }} }} }});
    buildHeatmap('lenderHeatmap', 'lender', {json.dumps(lender_data["heatmap_entities"])}, {json.dumps(lender_data["heatmap_topics"])}, {json.dumps(lender_data["heatmap_data"])}, {json.dumps(lender_data["heatmap_topic_polarities"])});
  }}
  if (layer === 'borrower') {{
    chartInstances['borrowerChart'] = new Chart(document.getElementById('borrowerChart'), {{ type:'radar', data:{{ labels:{borrower_radar_labels}, datasets:{borrower_radar_datasets} }}, options:{{ responsive:true, plugins:{{legend:{{position:'bottom',labels:{{boxWidth:12}}}}}}, scales:{{ r:{{beginAtZero:true, grid:{{color:'#21262d'}}, angleLines:{{color:'#21262d'}}, pointLabels:{{font:{{size:10}},color:'#c9d1d9'}}}} }} }} }});
    const bScores = {json.dumps(borrower_data["scores"])};
    const bLabels = {json.dumps(borrower_data["labels"])};
    const bScoreCanvas = document.getElementById('borrowerScoreChart');
    const bScoreWrap = bScoreCanvas && bScoreCanvas.closest('.borrower-score-chart-wrap');
    if (bScoreWrap) {{ bScoreWrap.style.height = Math.max(320, bLabels.length * 28) + 'px'; }}
    chartInstances['borrowerScoreChart'] = new Chart(bScoreCanvas, {{ type:'bar', data:{{ labels:bLabels, datasets:[{{ label:'Stress Score', data:bScores, backgroundColor:bScores.map(v=>v>=70?'#FF6B6B':v>=50?'#FFD93D':'#00d4aa'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}, ticks:{{autoSkip:false, maxRotation:0}}}} }} }} }});
    buildHeatmap('borrowerHeatmap', 'borrower', {json.dumps(borrower_data["heatmap_entities"])}, {json.dumps(borrower_data["heatmap_topics"])}, {json.dumps(borrower_data["heatmap_data"])}, {json.dumps(borrower_data["heatmap_topic_polarities"])});
  }}
{bank_chart_block}
}}

/* ── API Key helpers ─────────────────────────────────── */
const API_KEY_STORAGE = 'bigdata_api_key';
function getUserApiKey() {{ return localStorage.getItem(API_KEY_STORAGE) || ''; }}
function setUserApiKey(k) {{ localStorage.setItem(API_KEY_STORAGE, k); }}
function hasUserApiKey() {{ return !!getUserApiKey(); }}

async function apiRequest(url, opts = {{}}) {{
  const key = getUserApiKey();
  if (!opts.headers) opts.headers = {{}};
  if (key) opts.headers['X-API-KEY'] = key;
  const resp = await fetch(url, opts);
  if (resp.status === 401) {{
    openSettings();
    throw new Error('API key required');
  }}
  return resp;
}}

let apiKeyGateActive = false;

function openSettings(isGate) {{
  apiKeyGateActive = !!isGate;
  const ov = document.getElementById('settingsOverlay');
  const inp = document.getElementById('settingsApiKey');
  const desc = document.getElementById('settingsApiKeyDescription');
  const cancel = ov.querySelector('.btn-close-settings');
  const status = document.getElementById('apiKeyStatus');
  inp.value = getUserApiKey();
  if (isGate) {{
    desc.textContent = 'Enter your Bigdata API key to continue. The dashboard will be available after you save.';
    cancel.style.display = 'none';
    status.textContent = 'Enter your key below to get started.';
  }} else {{
    desc.textContent = 'Enter or update your Bigdata API key for authenticated requests.';
    cancel.style.display = '';
    status.textContent = hasUserApiKey() ? '\\u2713 Custom API key is set' : 'No key set.';
  }}
  ov.classList.add('visible');
}}

function closeSettings() {{
  if (apiKeyGateActive) return;
  document.getElementById('settingsOverlay').classList.remove('visible');
}}

function saveApiKey() {{
  const key = document.getElementById('settingsApiKey').value.trim();
  if (!key) return;
  setUserApiKey(key);
  const wasGate = apiKeyGateActive;
  apiKeyGateActive = false;
  document.getElementById('settingsOverlay').classList.remove('visible');
  if (wasGate) {{
    document.getElementById('appContent').style.display = 'flex';
  }}
}}

function clearApiKey() {{
  document.getElementById('settingsApiKey').value = '';
  localStorage.removeItem(API_KEY_STORAGE);
  document.getElementById('apiKeyStatus').textContent = 'Key cleared.';
}}

function toggleApiKeyVisibility() {{
  const inp = document.getElementById('settingsApiKey');
  inp.type = inp.type === 'password' ? 'text' : 'password';
}}

document.getElementById('settingsOverlay').addEventListener('click', function(e) {{
  if (e.target === this && !apiKeyGateActive) closeSettings();
}});

/* ── Launch gate ─────────────────────────────────────── */
(function() {{
  if (!hasUserApiKey()) {{
    document.getElementById('appContent').style.display = 'none';
    openSettings(true);
  }} else {{
    document.getElementById('appContent').style.display = 'flex';
  }}
}})();

/* ── Custom entity state ─────────────────────────────── */
let currentCustomLayer = null;
let confirmedEntities = [];
const defaultLayerData = {{
  lender: {{ labels:{json.dumps(lender_data["labels"])}, scores:{json.dumps(lender_data["scores"])}, score_col:'{lender_data["score_col"]}', heatmap_entities:{json.dumps(lender_data["heatmap_entities"])}, heatmap_topics:{json.dumps(lender_data["heatmap_topics"])}, heatmap_data:{json.dumps(lender_data["heatmap_data"])}, heatmap_topic_polarities:{json.dumps(lender_data["heatmap_topic_polarities"])}, audit_docs:{json.dumps(lender_data["audit_docs"], default=str)}, entity_count:{lender_data["entity_count"]}, topic_count:{lender_data["topic_count"]}, radar_datasets:{json.dumps(lender_data.get("radar_datasets", []))}, theme_topics:{json.dumps(lender_data["theme_topics"])} }},
  borrower: {{ labels:{json.dumps(borrower_data["labels"])}, scores:{json.dumps(borrower_data["scores"])}, score_col:'{borrower_data["score_col"]}', heatmap_entities:{json.dumps(borrower_data["heatmap_entities"])}, heatmap_topics:{json.dumps(borrower_data["heatmap_topics"])}, heatmap_data:{json.dumps(borrower_data["heatmap_data"])}, heatmap_topic_polarities:{json.dumps(borrower_data["heatmap_topic_polarities"])}, audit_docs:{json.dumps(borrower_data["audit_docs"], default=str)}, entity_count:{borrower_data["entity_count"]}, topic_count:{borrower_data["topic_count"]}, radar_datasets:{json.dumps(borrower_data.get("radar_datasets", []))}, theme_topics:{json.dumps(borrower_data["theme_topics"])} }}
}};
let isCustom = {{ lender: false, borrower: false }};

function openCustomEntityModal() {{
  const layer = document.querySelector('.nav-item.active')?.getAttribute('data-layer');
  if (!layer || layer === 'overview') return;
  currentCustomLayer = layer;
  document.getElementById('entityLayerLabel').textContent = layer;
  document.getElementById('entityNamesInput').value = '';
  document.getElementById('entityStepInput').classList.add('active');
  document.getElementById('entityStepConfirm').classList.remove('active');
  document.getElementById('entityOverlay').classList.add('visible');
}}

function closeEntityOverlay() {{
  document.getElementById('entityOverlay').classList.remove('visible');
}}

function backToEntityInput() {{
  document.getElementById('entityStepConfirm').classList.remove('active');
  document.getElementById('entityStepInput').classList.add('active');
}}

function parseCompanyNames(raw) {{
  const t = raw.trim();
  if (!t) return [];
  const lines = t.split(/\\r?\\n/).map(s => s.trim()).filter(Boolean);
  if (lines.length > 1) return lines;
  return t.split(',').map(s => s.trim()).filter(Boolean);
}}

async function lookupEntities() {{
  const raw = document.getElementById('entityNamesInput').value;
  const names = parseCompanyNames(raw);
  if (!names.length) return;
  const btn = document.querySelector('#entityStepInput .btn-save-key');
  btn.textContent = 'Looking up...';
  btn.disabled = true;
  try {{
    const resp = await apiRequest('/api/company-lookup', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify({{ names, layer: currentCustomLayer }})
    }});
    const data = await resp.json();
    confirmedEntities = [];
    data.results.forEach(r => {{
      if (r.matches && r.matches.length > 0) {{
        const m = r.matches[0];
        confirmedEntities.push({{
          input_name: r.input_name,
          name: m.name || r.input_name,
          rp_entity_id: m.id || '',
          country: m.country || '',
          sector: m.sector || '',
          ticker: (m.listing_values && m.listing_values[0]) ? m.listing_values[0].split(':')[1] : null,
          type: m.type || '',
        }});
      }} else {{
        confirmedEntities.push({{ input_name: r.input_name, name: r.input_name, rp_entity_id: '', country: '', sector: '', ticker: null, type: 'Unknown', error: r.error || 'Not found' }});
      }}
    }});
    renderEntityTable();
    document.getElementById('entityStepInput').classList.remove('active');
    document.getElementById('entityStepConfirm').classList.add('active');
  }} catch (e) {{
    alert('Lookup failed: ' + e.message);
  }} finally {{
    btn.textContent = 'Look Up';
    btn.disabled = false;
  }}
}}

function renderEntityTable() {{
  let h = '<table class="entity-table"><thead><tr><th>Your Input</th><th>Bigdata Name</th><th>Entity ID</th><th>Country</th><th>Sector</th><th></th></tr></thead><tbody>';
  confirmedEntities.forEach((e, i) => {{
    const errCls = e.error ? 'style="color:#FF6B6B"' : '';
    h += '<tr><td>'+escapeHtml(e.input_name)+'</td><td '+errCls+'>'+escapeHtml(e.name)+(e.error ? ' ('+escapeHtml(e.error)+')' : '')+'</td><td>'+escapeHtml(e.rp_entity_id)+'</td><td>'+escapeHtml(e.country)+'</td><td>'+escapeHtml(e.sector)+'</td><td><button class="btn-remove-entity" onclick="removeEntity('+i+')">Remove</button></td></tr>';
  }});
  h += '</tbody></table>';
  document.getElementById('entityTableWrap').innerHTML = h;
}}

function removeEntity(idx) {{
  confirmedEntities.splice(idx, 1);
  renderEntityTable();
}}

async function confirmEntities() {{
  const valid = confirmedEntities.filter(e => !e.error);
  if (!valid.length) {{ alert('No valid entities to run.'); return; }}
  closeEntityOverlay();
  const payload = {{
    entities: valid.map(e => ({{
      name: e.name,
      rp_entity_id: e.rp_entity_id || null,
      layer: currentCustomLayer,
      ticker: e.ticker || null
    }}))
  }};
  document.getElementById('pipelineOverlay').classList.add('visible');
  document.getElementById('pipelineProgress').textContent = 'Starting...';
  document.getElementById('pipelineElapsed').textContent = '';
  try {{
    const runResp = await apiRequest('/api/pipeline/run', {{
      method: 'POST',
      headers: {{ 'Content-Type': 'application/json' }},
      body: JSON.stringify(payload)
    }});
    const {{ job_id }} = await runResp.json();
    await pollPipeline(job_id);
  }} catch (e) {{
    document.getElementById('pipelineOverlay').classList.remove('visible');
    alert('Pipeline failed: ' + e.message);
  }}
}}

async function pollPipeline(jobId) {{
  while (true) {{
    await new Promise(r => setTimeout(r, 2000));
    try {{
      const resp = await apiRequest('/api/pipeline/status/' + jobId);
      const st = await resp.json();
      document.getElementById('pipelineProgress').textContent = st.progress || st.status;
      document.getElementById('pipelineElapsed').textContent = st.elapsed_ms ? (st.elapsed_ms / 1000).toFixed(1) + 's elapsed' : '';
      if (st.status === 'complete') {{
        const dataResp = await apiRequest('/api/pipeline/data/' + jobId);
        const layerData = await dataResp.json();
        document.getElementById('pipelineOverlay').classList.remove('visible');
        updateLayerData(currentCustomLayer, layerData);
        return;
      }}
      if (st.status === 'failed') {{
        document.getElementById('pipelineOverlay').classList.remove('visible');
        alert('Pipeline failed: ' + (st.error || 'Unknown error'));
        return;
      }}
    }} catch (e) {{
      document.getElementById('pipelineOverlay').classList.remove('visible');
      alert('Polling error: ' + e.message);
      return;
    }}
  }}
}}

/* ── Chart instance tracking for re-rendering ────────── */
const chartInstances = {{}};

function destroyCharts(layer) {{
  const ids = {{
    lender: ['lenderRadarChart', 'lenderChart'],
    borrower: ['borrowerChart', 'borrowerScoreChart'],
    bank: ['bankChart'],
  }};
  (ids[layer] || []).forEach(cid => {{
    const el = document.getElementById(cid);
    if (!el) return;
    const existing = typeof Chart !== 'undefined' && Chart.getChart ? Chart.getChart(el) : null;
    if (existing) existing.destroy();
    delete chartInstances[cid];
  }});
}}

function updateLayerData(layer, data) {{
  isCustom[layer] = true;
  auditData[layer] = data.audit_docs || [];
  auditsInitialized[layer] = false;

  destroyCharts(layer);
  chartsInitialized[layer] = false;

  pageBadges[layer] = data.entity_count + ' entities (custom) &middot; ' + data.topic_count + ' topics';
  const activeLyr = document.querySelector('.nav-item.active')?.getAttribute('data-layer');
  if (activeLyr === layer) document.getElementById('pageBadge').innerHTML = pageBadges[layer];

  /* Rebuild stat cards */
  const page = document.getElementById('page-' + layer);
  if (page) {{
    const statVals = page.querySelectorAll('.stat-val');
    if (statVals.length >= 4) {{
      statVals[0].textContent = data.entity_count;
      statVals[1].textContent = data.topic_count;
      const scores = data.scores || [];
      statVals[2].textContent = scores.length ? Math.max(...scores).toFixed(1) : '0.0';
      statVals[3].textContent = scores.length ? Math.min(...scores).toFixed(1) : '0.0';
    }}
  }}

  /* Rebuild charts */
  if (layer === 'lender') {{
    const radarLabels = data.labels || [];
    const radarDS = data.radar_datasets || [];
    chartInstances['lenderRadarChart'] = new Chart(document.getElementById('lenderRadarChart'), {{ type:'radar', data:{{ labels:radarLabels, datasets:radarDS }}, options:{{ responsive:true, plugins:{{legend:{{position:'bottom',labels:{{boxWidth:12}}}}}}, scales:{{ r:{{beginAtZero:true, grid:{{color:'#21262d'}}, angleLines:{{color:'#21262d'}}, pointLabels:{{font:{{size:10}},color:'#c9d1d9'}}}} }} }} }});
    const s = data.scores;
    const lLabels = data.labels;
    const lScoreCanvas = document.getElementById('lenderChart');
    const lScoreWrap = lScoreCanvas && lScoreCanvas.closest('.lender-score-chart-wrap');
    if (lScoreWrap) lScoreWrap.style.height = Math.max(320, lLabels.length * 28) + 'px';
    chartInstances['lenderChart'] = new Chart(lScoreCanvas, {{ type:'bar', data:{{ labels:lLabels, datasets:[{{ label:'Terms Power Score', data:s, backgroundColor:s.map(v=>v>50?'#00d4aa':v>30?'#FFD93D':'#FF6B6B'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}, ticks:{{autoSkip:false, maxRotation:0}}}} }} }} }});
    buildHeatmap('lenderHeatmap', 'lender', data.heatmap_entities, data.heatmap_topics, data.heatmap_data, data.heatmap_topic_polarities);
  }}
  if (layer === 'borrower') {{
    const radarLabels = data.labels || [];
    const radarDS = data.radar_datasets || [];
    chartInstances['borrowerChart'] = new Chart(document.getElementById('borrowerChart'), {{ type:'radar', data:{{ labels:radarLabels, datasets:radarDS }}, options:{{ responsive:true, plugins:{{legend:{{position:'bottom',labels:{{boxWidth:12}}}}}}, scales:{{ r:{{beginAtZero:true, grid:{{color:'#21262d'}}, angleLines:{{color:'#21262d'}}, pointLabels:{{font:{{size:10}},color:'#c9d1d9'}}}} }} }} }});
    const bScores = data.scores;
    const bLabels = data.labels;
    const bScoreCanvas = document.getElementById('borrowerScoreChart');
    const bScoreWrap = bScoreCanvas && bScoreCanvas.closest('.borrower-score-chart-wrap');
    if (bScoreWrap) bScoreWrap.style.height = Math.max(320, bLabels.length * 28) + 'px';
    chartInstances['borrowerScoreChart'] = new Chart(bScoreCanvas, {{ type:'bar', data:{{ labels:bLabels, datasets:[{{ label:'Stress Score', data:bScores, backgroundColor:bScores.map(v=>v>=70?'#FF6B6B':v>=50?'#FFD93D':'#00d4aa'), borderRadius:4 }}] }}, options:{{ indexAxis:'y', responsive:true, maintainAspectRatio:false, plugins:{{legend:{{display:false}}}}, scales:{{ x:{{min:0,max:100,grid:{{color:'#21262d'}}}}, y:{{grid:{{display:false}}, ticks:{{autoSkip:false, maxRotation:0}}}} }} }} }});
    buildHeatmap('borrowerHeatmap', 'borrower', data.heatmap_entities, data.heatmap_topics, data.heatmap_data, data.heatmap_topic_polarities);
  }}
  chartsInitialized[layer] = true;
  initAudit(layer);

  /* Show reset button, hide customize */
  document.getElementById('btnResetCustom').style.display = 'inline-block';
}}

function resetToDefault() {{
  const layer = document.querySelector('.nav-item.active')?.getAttribute('data-layer');
  if (!layer || !isCustom[layer]) return;
  const data = defaultLayerData[layer];
  if (!data) return;
  isCustom[layer] = false;
  auditData[layer] = data.audit_docs || [];
  auditsInitialized[layer] = false;
  destroyCharts(layer);
  chartsInitialized[layer] = false;
  pageBadges[layer] = data.entity_count + ' entities &middot; ' + data.topic_count + ' topics';
  document.getElementById('pageBadge').innerHTML = pageBadges[layer];
  initCharts(layer);
  initAudit(layer);
  document.getElementById('btnResetCustom').style.display = 'none';
}}

/* ── Extend switchLayer to show/hide customize button ── */
const _origSwitchLayer = switchLayer;
switchLayer = function(layer) {{
  _origSwitchLayer(layer);
  const btn = document.getElementById('btnCustomize');
  const rst = document.getElementById('btnResetCustom');
  if (layer === 'overview') {{
    btn.style.display = 'none';
    rst.style.display = 'none';
  }} else {{
    btn.style.display = 'inline-block';
    rst.style.display = isCustom[layer] ? 'inline-block' : 'none';
  }}
}};

</script>
</body>
</html>"""


def generate_reports(df: pd.DataFrame | None = None) -> None:
    """Generate both Excel and HTML reports from scores."""
    if df is None:
        if not SCORES_CSV.exists():
            raise FileNotFoundError(
                f"Scores file not found at {SCORES_CSV}. Run scorer first."
            )
        df = pd.read_csv(SCORES_CSV)

    console.rule("[bold cyan]Report Generation")
    excel_path = generate_excel(df)
    html_path = generate_html_dashboard(df)
    console.print(f"[green]Excel:[/green] {excel_path}")
    console.print(f"[green]HTML:[/green]  {html_path} (publish `dist/` via GitHub Pages — see README)")


if __name__ == "__main__":
    generate_reports()
